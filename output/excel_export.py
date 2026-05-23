"""
CSV + XLSX export with a "Top Leads" highlight sheet.

The XLSX file gets two sheets:
- "Leads"     — every row, sorted by score
- "Top 50"    — the 50 highest-scoring facilities with conditional formatting
"""
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from ..config import CONFIG


# Columns to surface in the output, in priority order
DISPLAY_COLUMNS = [
    "score_total",
    "name",
    "address", "city", "state", "zip",
    "beds", "ownership_type", "cert_date",
    "true_owner", "legal_owner",
    "owner_name", "owner_title", "owner_email",
    "phone_e164", "phone",
    "website", "operator_domain", "operator_linkedin",
    "operator_employee_count", "operator_founded_year",
    "rating", "rating_count",
    "google_url",
    "score_size_fit", "score_age_of_ownership", "score_operator_independence",
    "score_operator_age", "score_occupancy_proxy", "score_quality_risk",
    "score_geo_fit",
    "source",
]


def export_leads(df: pd.DataFrame, prefix: str = "senior_housing_leads") -> Path:
    """Write CSV + XLSX to the configured output dir; return the XLSX path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir: Path = CONFIG.output_dir
    csv_path = out_dir / f"{prefix}_{ts}.csv"
    xlsx_path = out_dir / f"{prefix}_{ts}.xlsx"

    cols = [c for c in DISPLAY_COLUMNS if c in df.columns]
    extra = [c for c in df.columns if c not in cols]
    df = df[cols + extra]

    df.to_csv(csv_path, index=False)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Leads", index=False)
        top = df.head(50)
        top.to_excel(writer, sheet_name="Top 50", index=False)

        # Conditional formatting on the score column for the Top 50 sheet
        ws = writer.sheets["Top 50"]
        if "score_total" in top.columns:
            col_idx = list(top.columns).index("score_total") + 1
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}2:{col_letter}{len(top) + 1}"
            rule = ColorScaleRule(
                start_type="num", start_value=0, start_color="FFF8696B",
                mid_type="num",   mid_value=50, mid_color="FFFFEB84",
                end_type="num",   end_value=100, end_color="FF63BE7B",
            )
            ws.conditional_formatting.add(rng, rule)

        # Reasonable column widths
        for sheet in writer.sheets.values():
            for idx, col in enumerate(df.columns, start=1):
                width = min(40, max(12, df[col].astype(str).str.len().max() if not df[col].empty else 12))
                sheet.column_dimensions[get_column_letter(idx)].width = width

    return xlsx_path
